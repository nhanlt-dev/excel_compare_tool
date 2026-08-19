import pandas as pd
from openpyxl import load_workbook
from utils.logger import get_logger

logger = get_logger()

def load_excel_sheets(path):
    """Return workbook sheet names without loading cell data."""
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def detect_header_row(path, sheet_name, scan_rows=30):
    """Choose the most header-like row: many non-empty, mostly text, unique cells."""
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb[sheet_name]
        best_row, best_score = 1, float("-inf")
        for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1):
            values = [v for v in row if v not in (None, "")]
            if not values:
                continue
            texts = [str(v).strip() for v in values]
            text_count = sum(not str(v).replace(".", "", 1).isdigit() for v in values)
            unique_ratio = len(set(texts)) / len(texts)
            # Prefer several distinct text labels; penalize title rows with one long cell.
            score = len(values) * 4 + text_count * 2 + unique_ratio - (8 if len(values) == 1 else 0)
            if score > best_score:
                best_row, best_score = row_no, score
        return best_row
    finally:
        wb.close()


def load_excel_columns(path, sheet_name=0, header_row=1):
    # read only headers
    try:
        df = pd.read_excel(
            path, sheet_name=sheet_name, header=header_row - 1,
            nrows=0, engine="openpyxl"
        )
        return [str(c) for c in df.columns]
    except Exception as e:
        logger.error(f"Failed to load Excel columns from {path}: {e}")
        raise e

def read_table(path):
    # read all as object to preserve values
    try:
        df = pd.read_excel(path, dtype=object, engine="openpyxl")
        logger.info(f"Loaded table from {path}: {len(df)} rows, {len(df.columns)} columns")
        return df
    except Exception as e:
        logger.error(f"Failed to read table from {path}: {e}")
        raise e

def check_duplicate_keys(df, key_column):
    """
    Check for duplicate values in a key column.

    Args:
        df: pandas DataFrame
        key_column: name of the column to check

    Returns:
        tuple: (has_duplicates, duplicate_count, duplicate_values)
               duplicate_count is the number of duplicate *occurrences* (not unique values)
               duplicate_values is a list of the duplicated values (unique values that appear >1 time)
    """
    if key_column not in df.columns:
        logger.warning(f"Key column '{key_column}' not found in DataFrame")
        return False, 0, []

    # Drop NaN values for duplicate checking
    keys = df[key_column].dropna()
    duplicates = keys[keys.duplicated(keep=False)]

    if len(duplicates) > 0:
        dup_values = duplicates.unique().tolist()
        # Count how many extra rows are caused by duplicates (total rows - unique rows)
        dup_count = len(duplicates) - len(dup_values)
        logger.warning(f"Found {len(dup_values)} unique duplicate values in column '{key_column}' ({dup_count} extra rows)")
        return True, dup_count, dup_values
    return False, 0, []
