import os
import re
import shutil
import unicodedata
from difflib import SequenceMatcher
from datetime import datetime
from collections import Counter
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from utils.logger import get_logger

logger = get_logger()
NUMBER_RE = re.compile(r"^[+-]?\d+(?:[.,]\d+)?(?:[eE][+-]?\d+)?$")


def normalize_value(value, case_sensitive=False, remove_accents=False):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    numeric = text.replace("\u2212", "-")
    if NUMBER_RE.match(numeric):
        try:
            number = Decimal(numeric.replace(",", "."))
            return format(number.normalize(), "f")
        except InvalidOperation:
            pass
    if not case_sensitive:
        text = text.casefold()
    if remove_accents:
        text = "".join(
            c for c in unicodedata.normalize("NFKD", text)
            if not unicodedata.combining(c)
        )
    return text


def _header_map(ws, header_row):
    result = {}
    duplicates = set()
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name in result:
            duplicates.add(name)
        else:
            result[name] = cell.column
    if duplicates:
        raise ValueError("Tiêu đề cột bị trùng: " + ", ".join(sorted(duplicates)))
    return result


def _make_key(ws, row, columns, case_sensitive, remove_accents):
    values = tuple(
        normalize_value(ws.cell(row, col).value, case_sensitive, remove_accents)
        for col in columns
    )
    if all(value == "" for value in values):
        return None
    return values


def _make_key_from_values(values, zero_based_columns, case_sensitive, remove_accents):
    key = tuple(
        normalize_value(values[col], case_sensitive, remove_accents)
        for col in zero_based_columns
    )
    if all(value == "" for value in key):
        return None
    return key


def _report_progress(callback, phase, done, total):
    if callback:
        callback(phase, done, max(1, total))


def lookup_and_update(
    source_path, target_path, output_path, source_sheet, target_sheet,
    key_pairs, value_pairs, source_header_row=1, target_header_row=1,
    duplicate_policy="first", missing_policy="keep", overwrite_policy="all",
    missing_text="KHÔNG TÌM THẤY", case_sensitive=False,
    remove_accents=False, progress_callback=None, stop_event=None,
    create_audit=True, highlight_updates=False
):
    """Lookup composite keys in source and update a copied target workbook."""
    if not key_pairs:
        raise ValueError("Cần ít nhất một cặp khóa đối chiếu.")
    if not value_pairs:
        raise ValueError("Cần ít nhất một cặp cột lấy/ghi dữ liệu.")
    if os.path.abspath(target_path) == os.path.abspath(output_path):
        raise ValueError("File kết quả phải khác file đích gốc.")

    shutil.copy2(target_path, output_path)
    source_wb = load_workbook(source_path, read_only=True, data_only=False)
    target_wb = load_workbook(output_path, data_only=False)
    try:
        if source_sheet not in source_wb.sheetnames:
            raise ValueError(f"Không tìm thấy Sheet nguồn '{source_sheet}'.")
        if target_sheet not in target_wb.sheetnames:
            raise ValueError(f"Không tìm thấy Sheet đích '{target_sheet}'.")
        src = source_wb[source_sheet]
        dst = target_wb[target_sheet]

        audit_ws = missing_ws = duplicate_ws = summary_ws = suggestion_ws = None
        report_names = ("TONG_HOP", "NHAT_KY_CAP_NHAT", "KHONG_TIM_THAY", "KHOA_NGUON_TRUNG", "GOI_Y_GAN_DUNG")
        if not create_audit:
            for name in report_names:
                if name in target_wb.sheetnames and name != target_sheet:
                    target_wb.remove(target_wb[name])
        if create_audit:
            for name in report_names:
                if name in target_wb.sheetnames:
                    target_wb.remove(target_wb[name])
            audit_ws = target_wb.create_sheet("NHAT_KY_CAP_NHAT")
            audit_ws.append(["Sheet đích", "Dòng", "Khóa", "Cột", "Giá trị cũ", "Giá trị mới", "Kiểu nền cũ", "Màu nền cũ"])
            missing_ws = target_wb.create_sheet("KHONG_TIM_THAY")
            missing_ws.append(["Sheet đích", "Dòng", "Khóa"])
            duplicate_ws = target_wb.create_sheet("KHOA_NGUON_TRUNG")
            duplicate_ws.append(["Khóa nguồn", "Số lần xuất hiện"])
            suggestion_ws = target_wb.create_sheet("GOI_Y_GAN_DUNG")
            suggestion_ws.append(["Sheet đích", "Dòng", "Khóa không tìm thấy", "Gợi ý khóa nguồn", "Độ tương đồng (%)"])
            summary_ws = target_wb.create_sheet("TONG_HOP", 0)
            summary_ws.append(["EXCEL COMPARE PRO - BÁO CÁO TỔNG HỢP", ""])
            summary_ws.append(["Thời gian chạy", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            summary_ws.append(["File nguồn", os.path.basename(source_path)])
            summary_ws.append(["Sheet nguồn", source_sheet])
            summary_ws.append(["File đích", os.path.basename(target_path)])
            summary_ws.append(["Sheet đích", target_sheet])
            summary_ws.append(["Cặp khóa", "; ".join(f"{a} → {b}" for a, b in key_pairs)])
            summary_ws.append(["Cột lấy → ghi", "; ".join(f"{a} → {b}" for a, b in value_pairs)])
            for sheet in (audit_ws, missing_ws, duplicate_ws, suggestion_ws):
                for cell in sheet[1]: cell.font = Font(bold=True)
                sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
            summary_ws["A1"].font = Font(bold=True, size=14)
        src_headers = _header_map(src, source_header_row)
        dst_headers = _header_map(dst, target_header_row)

        missing_src = [c for pair in key_pairs + value_pairs for c in [pair[0]] if c not in src_headers]
        missing_dst = [c for pair in key_pairs + value_pairs for c in [pair[1]] if c not in dst_headers]
        if missing_src:
            raise ValueError("Thiếu cột nguồn: " + ", ".join(sorted(set(missing_src))))
        if missing_dst:
            raise ValueError("Thiếu cột đích: " + ", ".join(sorted(set(missing_dst))))

        src_key_cols = [src_headers[a] for a, _ in key_pairs]
        dst_key_cols = [dst_headers[b] for _, b in key_pairs]
        src_value_cols = [src_headers[a] for a, _ in value_pairs]
        dst_value_cols = [dst_headers[b] for _, b in value_pairs]

        lookup = {}
        key_counts = Counter()
        # Some workbooks omit the worksheet dimension metadata, so max_row can
        # be None in read-only mode. Progress remains valid with a safe floor.
        source_total = max(1, (src.max_row or 0) - source_header_row)
        source_max_col = max(src_key_cols + src_value_cols)
        src_key_idx = [col - 1 for col in src_key_cols]
        src_value_idx = [col - 1 for col in src_value_cols]

        # Streaming read is essential here. Random ws.cell() access on a
        # read-only worksheet can repeatedly scan the underlying XML.
        source_rows = src.iter_rows(
            min_row=source_header_row + 1,
            max_col=source_max_col,
            values_only=True
        )
        source_processed = 0
        for position, values in enumerate(source_rows, start=1):
            source_processed = position
            if stop_event is not None and stop_event.is_set():
                raise InterruptedError("Đã dừng theo yêu cầu người dùng.")
            key = _make_key_from_values(
                values, src_key_idx, case_sensitive, remove_accents
            )
            if key is None:
                if position % 250 == 0:
                    _report_progress(progress_callback, "source", position, max(source_total, position))
                continue
            key_counts[key] += 1
            result_values = tuple(values[col] for col in src_value_idx)
            if duplicate_policy == "last" or key not in lookup:
                lookup[key] = result_values
            if position % 250 == 0:
                _report_progress(progress_callback, "source", position, max(source_total, position))
        source_total = max(source_total, source_processed)
        _report_progress(progress_callback, "source", source_total, source_total)

        duplicate_keys = {key for key, count in key_counts.items() if count > 1}
        # Index candidates by all key components except the last. Fuzzy matching
        # is report-only and never writes data automatically.
        fuzzy_groups = {}
        for source_key in lookup:
            fuzzy_groups.setdefault(source_key[:-1], []).append(source_key)
        if duplicate_ws:
            for key in sorted(duplicate_keys, key=str):
                duplicate_ws.append([" | ".join(key), key_counts[key]])
        if duplicate_policy == "error" and duplicate_keys:
            raise ValueError(f"Nguồn có {len(duplicate_keys)} khóa trùng. Hãy chọn cách xử lý khóa trùng.")

        stats = {"total": 0, "matched": 0, "missing": 0, "updated_cells": 0,
                 "skipped_cells": 0, "duplicate_keys": len(duplicate_keys)}
        suggestion_count = 0
        total_rows = max(0, dst.max_row - target_header_row)
        target_max_col = max(dst_key_cols + dst_value_cols)
        dst_key_idx = [col - 1 for col in dst_key_cols]
        target_rows = dst.iter_rows(
            min_row=target_header_row + 1,
            max_col=target_max_col
        )
        for position, row_cells in enumerate(target_rows, start=1):
            if stop_event is not None and stop_event.is_set():
                raise InterruptedError("Đã dừng theo yêu cầu người dùng.")
            row_values = tuple(cell.value for cell in row_cells)
            key = _make_key_from_values(
                row_values, dst_key_idx, case_sensitive, remove_accents
            )
            if key is None:
                if position % 100 == 0:
                    _report_progress(progress_callback, "target", position, total_rows)
                continue
            stats["total"] += 1
            values = lookup.get(key)
            if values is None:
                stats["missing"] += 1
                if missing_ws:
                    missing_ws.append([target_sheet, target_header_row + position, " | ".join(key)])
                if suggestion_ws and key and suggestion_count < 1000:
                    candidates = fuzzy_groups.get(key[:-1], [])
                    best = None; best_score = 0.0
                    for candidate in candidates[:500]:
                        score = SequenceMatcher(None, key[-1], candidate[-1]).ratio()
                        if score > best_score:
                            best, best_score = candidate, score
                    if best is not None and best_score >= 0.65:
                        suggestion_ws.append([
                            target_sheet, target_header_row + position, " | ".join(key),
                            " | ".join(best), round(best_score * 100, 1)
                        ])
                        suggestion_count += 1
                for col in dst_value_cols:
                    cell = row_cells[col - 1]
                    if missing_policy == "clear":
                        cell.value = None
                    elif missing_policy == "text":
                        cell.value = missing_text
                if position % 100 == 0:
                    _report_progress(progress_callback, "target", position, total_rows)
                continue

            stats["matched"] += 1
            for col, value in zip(dst_value_cols, values):
                cell = row_cells[col - 1]
                if overwrite_policy == "blank_only" and cell.value not in (None, ""):
                    stats["skipped_cells"] += 1
                    continue
                old_value = cell.value
                old_fill_type = cell.fill.fill_type
                old_fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else None
                cell.value = value
                if highlight_updates:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                if audit_ws:
                    audit_ws.append([
                        target_sheet, target_header_row + position, " | ".join(key),
                        dst.cell(target_header_row, col).value, old_value, value,
                        old_fill_type, old_fill_color
                    ])
                stats["updated_cells"] += 1
            if position % 100 == 0:
                _report_progress(progress_callback, "target", position, total_rows)

        if summary_ws:
            summary_ws.append(["", ""])
            summary_ws.append(["CHỈ SỐ", "KẾT QUẢ"])
            summary_ws.append(["Tổng dòng đích có khóa", stats["total"]])
            summary_ws.append(["Dòng khớp", stats["matched"]])
            summary_ws.append(["Dòng không tìm thấy", stats["missing"]])
            summary_ws.append(["Ô đã cập nhật", stats["updated_cells"]])
            summary_ws.append(["Ô bỏ qua", stats["skipped_cells"]])
            summary_ws.append(["Khóa nguồn trùng", stats["duplicate_keys"]])
            summary_ws.append(["Gợi ý gần đúng", suggestion_count])
            match_rate = stats["matched"] / stats["total"] if stats["total"] else 0
            summary_ws.append(["Tỷ lệ khớp", match_rate])
            summary_ws["B18"].number_format = "0.0%"
            summary_ws.column_dimensions["A"].width = 34
            summary_ws.column_dimensions["B"].width = 70
            summary_ws.freeze_panes = "A2"
            for cell in summary_ws[10]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
        if create_audit:
            for sheet in (audit_ws, missing_ws, duplicate_ws, suggestion_ws):
                sheet.auto_filter.ref = sheet.dimensions
                for column in sheet.columns:
                    letter = column[0].column_letter
                    sample = [len(str(cell.value)) for cell in column[:200] if cell.value is not None]
                    sheet.column_dimensions[letter].width = min(45, max([12] + sample) + 2)
        _report_progress(progress_callback, "target", total_rows, total_rows)
        _report_progress(progress_callback, "saving", 0, 1)
        target_wb.save(output_path)
        _report_progress(progress_callback, "saving", 1, 1)
        logger.info("Lookup update completed: %s", stats)
        return stats
    except Exception:
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except OSError:
                pass
        raise
    finally:
        source_wb.close()
        target_wb.close()


def undo_from_audit(updated_path, output_path):
    """Restore old values recorded by NHAT_KY_CAP_NHAT into a new workbook."""
    if os.path.abspath(updated_path) == os.path.abspath(output_path):
        raise ValueError("File hoàn tác phải khác file đã cập nhật.")
    shutil.copy2(updated_path, output_path)
    wb = load_workbook(output_path, data_only=False)
    try:
        if "NHAT_KY_CAP_NHAT" not in wb.sheetnames:
            raise ValueError("File không có Sheet NHAT_KY_CAP_NHAT để hoàn tác.")
        log = wb["NHAT_KY_CAP_NHAT"]
        restored = 0
        # Reverse order also handles multiple writes to the same cell safely.
        records = list(log.iter_rows(min_row=2, values_only=True))
        for record in reversed(records):
            sheet_name, row, _key, column_name, old_value, _new_value = record[:6]
            old_fill_type = record[6] if len(record) > 6 else None
            old_fill_color = record[7] if len(record) > 7 else None
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header_map = _header_map(ws, 1)
            if column_name not in header_map:
                # Find the header on any early row for non-row-1 workbooks.
                for header_row in range(1, min(30, ws.max_row) + 1):
                    candidate = _header_map(ws, header_row)
                    if column_name in candidate:
                        header_map = candidate; break
            if column_name in header_map:
                cell = ws.cell(int(row), header_map[column_name])
                cell.value = old_value
                if old_fill_type and old_fill_color:
                    cell.fill = PatternFill(old_fill_type, fgColor=old_fill_color)
                else:
                    cell.fill = PatternFill(fill_type=None)
                restored += 1
        wb.save(output_path)
        return restored
    except Exception:
        if os.path.exists(output_path):
            try: os.remove(output_path)
            except OSError: pass
        raise
    finally:
        wb.close()


def preview_lookup(source_path, target_path, source_sheet, target_sheet,
                   key_pairs, value_pairs, source_header_row=1,
                   target_header_row=1, case_sensitive=False,
                   remove_accents=False, limit=20):
    """Return a safe read-only preview without creating or changing a workbook."""
    source_wb = load_workbook(source_path, read_only=True, data_only=False)
    target_wb = load_workbook(target_path, read_only=True, data_only=False)
    try:
        src, dst = source_wb[source_sheet], target_wb[target_sheet]
        sh, dh = _header_map(src, source_header_row), _header_map(dst, target_header_row)
        src_key = [sh[a] - 1 for a, _ in key_pairs]
        dst_key = [dh[b] - 1 for _, b in key_pairs]
        src_val = [sh[a] - 1 for a, _ in value_pairs]
        dst_val = [dh[b] - 1 for _, b in value_pairs]
        max_src = max(src_key + src_val) + 1
        lookup = {}
        for values in src.iter_rows(min_row=source_header_row + 1, max_col=max_src, values_only=True):
            key = _make_key_from_values(values, src_key, case_sensitive, remove_accents)
            if key is not None and key not in lookup:
                lookup[key] = tuple(values[i] for i in src_val)
        result = []
        max_dst = max(dst_key + dst_val) + 1
        for excel_row, values in enumerate(dst.iter_rows(min_row=target_header_row + 1, max_col=max_dst, values_only=True), start=target_header_row + 1):
            key = _make_key_from_values(values, dst_key, case_sensitive, remove_accents)
            if key is None:
                continue
            new = lookup.get(key)
            result.append({
                "row": excel_row, "key": " | ".join(key),
                "old": tuple(values[i] for i in dst_val),
                "new": new, "status": "Khớp" if new is not None else "Không tìm thấy"
            })
            if len(result) >= limit:
                break
        return result
    finally:
        source_wb.close(); target_wb.close()
